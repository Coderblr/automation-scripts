import time
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService


BASE_HEADERS = [
    "Test_Case_ID",
    "Test_Case_Name",
    "Step_No",
    "Action",
    "Locator_Type",
    "Locator_Value",
    "Test_Data",
    "Expected_Result",
    "Run",
]

EXTRA_HEADERS = [
    "Priority",
    "Element_Type",
    "Status",
    "Actual_Result",
    "Comments",
]


def _log(logger: Optional[Callable[[str], None]], message: str) -> None:
    if logger:
        logger(message)


def _create_driver(browser: str):
    browser = (browser or "chrome").lower()

    if browser == "firefox":
        options = webdriver.FirefoxOptions()
        options.set_preference("signon.rememberSignons", False)
        options.set_preference("dom.webnotifications.enabled", False)
        try:
            service = FirefoxService(GeckoDriverManager().install())
            driver = webdriver.Firefox(service=service, options=options)
        except Exception:
            driver = webdriver.Firefox(options=options)
        return driver

    if browser == "edge":
        options = webdriver.EdgeOptions()
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-gpu")
        options.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
        options.add_experimental_option(
            "prefs",
            {
                "credentials_enable_service": False,
                "profile.password_manager_enabled": False,
                "profile.password_manager_leak_detection": False,
            },
        )
        try:
            service = EdgeService(EdgeChromiumDriverManager().install())
            driver = webdriver.Edge(service=service, options=options)
        except Exception:
            driver = webdriver.Edge(options=options)
        return driver

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-gpu")
    options.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
    options.add_experimental_option(
        "prefs",
        {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.password_manager_leak_detection": False,
        },
    )
    try:
        service = ChromeService(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
    except Exception:
        driver = webdriver.Chrome(options=options)
    return driver


def _by_to_locator_type(by_value: str) -> str:
    mapping = {
        By.ID: "id",
        By.NAME: "name",
        By.XPATH: "xpath",
        By.CSS_SELECTOR: "css",
        By.CLASS_NAME: "class",
        By.TAG_NAME: "tag",
        By.LINK_TEXT: "link_text",
        By.PARTIAL_LINK_TEXT: "partial_link_text",
    }
    return mapping.get(by_value, (by_value or "").lower().replace(" selector", ""))


def _find_first_visible(driver, selectors: List[Tuple[str, str]], timeout: int = 10):
    for by, value in selectors:
        try:
            wait = WebDriverWait(driver, timeout)
            el = wait.until(lambda d: d.find_element(by, value))
            if el and el.is_displayed():
                return el, by, value
        except Exception:
            continue
        try:
            el = driver.find_element(by, value)
            if el and el.is_displayed():
                return el, by, value
        except Exception:
            continue
    return None, "", ""


def _safe_login(driver, base_url: str, username: str, password: str, selectors: Dict[str, Tuple[str, str]]) -> bool:
    try:
        driver.get(base_url)
        time.sleep(2)
        u_el, _, _ = _find_first_visible(driver, [selectors["username"]], timeout=12)
        p_el, _, _ = _find_first_visible(driver, [selectors["password"]], timeout=12)
        s_el, _, _ = _find_first_visible(driver, [selectors["submit"]], timeout=12)
        if not u_el or not p_el or not s_el:
            return False
        u_el.clear()
        u_el.send_keys(username)
        p_el.clear()
        p_el.send_keys(password)
        s_el.click()
        time.sleep(3)
        return True
    except Exception:
        return False


def _discover_nav_candidates(driver, max_pages: int) -> List[Dict[str, str]]:
    nav_candidates = []
    seen = set()
    selectors = [
        "a[href]",
        "button",
        "[role='button']",
        "[class*='menu'] a",
        "[class*='menu'] button",
        "nav a",
        "nav button",
    ]
    combined = ",".join(selectors)
    elements = driver.find_elements(By.CSS_SELECTOR, combined)
    for el in elements:
        try:
            if not el.is_displayed():
                continue
            text = (el.text or "").strip()
            if len(text) < 2:
                continue
            key = text.lower()
            if key in seen:
                continue
            seen.add(key)
            locator_type, locator_value, etype = _best_locator(driver, el)
            nav_candidates.append(
                {
                    "text": text,
                    "locator_type": locator_type,
                    "locator_value": locator_value,
                    "etype": etype,
                }
            )
            if len(nav_candidates) >= max_pages:
                break
        except Exception:
            continue
    return nav_candidates


def _rowset_for_nav(
    tc_id: str,
    tc_name: str,
    login_locators: Optional[Dict[str, Tuple[str, str]]],
    nav: Dict[str, str],
    nav_url_fragment: str,
    heading_selector: str,
    heading_text: str,
) -> List[List[str]]:
    rows = [_step(tc_id, tc_name, 1, "navigate", test_data="${base_url}", priority="High")]
    step_no = 2

    if login_locators:
        u_by, u_val = login_locators["username"]
        p_by, p_val = login_locators["password"]
        s_by, s_val = login_locators["submit"]
        rows.extend(
            [
                _step(tc_id, tc_name, step_no, "type", _by_to_locator_type(u_by), u_val, "${username}", priority="High", element_type="text"),
                _step(tc_id, tc_name, step_no + 1, "type", _by_to_locator_type(p_by), p_val, "${password}", priority="High", element_type="password"),
                _step(tc_id, tc_name, step_no + 2, "click", _by_to_locator_type(s_by), s_val, priority="High", element_type="button"),
                _step(tc_id, tc_name, step_no + 3, "wait", test_data="3", priority="Medium"),
            ]
        )
        step_no += 4
    else:
        rows.append(_step(tc_id, tc_name, step_no, "wait", test_data="2", priority="Medium"))
        step_no += 1

    rows.extend(
        [
            _step(
                tc_id,
                tc_name,
                step_no,
                "click",
                nav["locator_type"],
                nav["locator_value"],
                priority=_to_priority(nav["etype"]),
                element_type=nav["etype"],
            ),
            _step(tc_id, tc_name, step_no + 1, "wait", test_data="2", priority="Medium"),
        ]
    )
    step_no += 2

    if nav_url_fragment:
        rows.append(_step(tc_id, tc_name, step_no, "verify_url", expected_result=nav_url_fragment, priority="Medium"))
        step_no += 1
    if heading_selector and heading_text:
        rows.append(
            _step(
                tc_id,
                tc_name,
                step_no,
                "verify_text",
                "css",
                heading_selector,
                expected_result=heading_text,
                priority="Medium",
            )
        )
        step_no += 1
    rows.append(_step(tc_id, tc_name, step_no, "take_screenshot", test_data=f"{tc_id}_{nav['text'].replace(' ', '_')}", priority="Low"))
    return rows


def _build_minimal_smoke_rows() -> List[List[str]]:
    return [
        _step("TC001", "Minimal Smoke - Open Application", 1, "navigate", test_data="${base_url}", priority="High"),
        _step("TC001", "Minimal Smoke - Open Application", 2, "wait", test_data="2", priority="Medium"),
        _step("TC001", "Minimal Smoke - Open Application", 3, "verify_url", expected_result="/", priority="Medium"),
        _step("TC001", "Minimal Smoke - Open Application", 4, "take_screenshot", test_data="TC001_smoke_home", priority="Low"),
        _step("TC002", "Minimal Smoke - Basic Interaction", 1, "navigate", test_data="${base_url}", priority="High"),
        _step("TC002", "Minimal Smoke - Basic Interaction", 2, "wait", test_data="2", priority="Medium"),
        _step("TC002", "Minimal Smoke - Basic Interaction", 3, "refresh", priority="Low"),
        _step("TC002", "Minimal Smoke - Basic Interaction", 4, "wait", test_data="1", priority="Low"),
        _step("TC002", "Minimal Smoke - Basic Interaction", 5, "take_screenshot", test_data="TC002_smoke_refresh", priority="Low"),
    ]


def _escape_xpath_text(text: str) -> str:
    if "'" not in text:
        return f"'{text}'"
    if '"' not in text:
        return f'"{text}"'
    parts = text.split("'")
    concat_parts = [f"'{parts[0]}'"]
    for part in parts[1:]:
        concat_parts.append("\"'\"")
        concat_parts.append(f"'{part}'")
    return "concat(" + ", ".join(concat_parts) + ")"


def _best_locator(driver, element) -> Tuple[str, str, str]:
    tag = (element.tag_name or "").lower()
    etype = (element.get_attribute("type") or "").lower() or tag
    element_id = (element.get_attribute("id") or "").strip()
    element_name = (element.get_attribute("name") or "").strip()
    data_testid = (element.get_attribute("data-testid") or "").strip()
    data_test = (element.get_attribute("data-test") or "").strip()
    aria_label = (element.get_attribute("aria-label") or "").strip()
    text = (element.text or "").strip()

    if element_id:
        return "id", element_id, etype
    if element_name:
        return "name", element_name, etype
    if data_testid:
        return "css", f"[data-testid='{data_testid}']", etype
    if data_test:
        return "css", f"[data-test='{data_test}']", etype
    if aria_label:
        return "css", f"[aria-label='{aria_label}']", etype

    if text and len(text) <= 40 and tag in ("a", "button", "span", "div"):
        escaped = _escape_xpath_text(text)
        return "xpath", f"//{tag}[normalize-space()={escaped}]", etype

    class_attr = (element.get_attribute("class") or "").strip().split()
    if class_attr:
        stable = [c for c in class_attr if len(c) > 2 and not c.startswith("ng-")]
        if stable:
            return "css", f"{tag}.{stable[0]}", etype

    try:
        xpath = driver.execute_script(
            """
            function absoluteXPath(element) {
              if (element.tagName.toLowerCase() === 'html') return '/html[1]';
              if (element === document.body) return '/html[1]/body[1]';
              let ix = 0;
              const siblings = element.parentNode.childNodes;
              for (let i = 0; i < siblings.length; i++) {
                const sibling = siblings[i];
                if (sibling === element) {
                  return absoluteXPath(element.parentNode) + '/' + element.tagName.toLowerCase() + '[' + (ix + 1) + ']';
                }
                if (sibling.nodeType === 1 && sibling.tagName === element.tagName) ix++;
              }
            }
            return absoluteXPath(arguments[0]);
            """,
            element,
        )
        if xpath:
            return "xpath", xpath, etype
    except Exception:
        pass

    return "css", tag, etype


def _extract_heading(driver) -> Tuple[str, str, str]:
    candidates = [
        (By.CSS_SELECTOR, "h1"),
        (By.CSS_SELECTOR, "h2"),
        (By.CSS_SELECTOR, "h3"),
        (By.CSS_SELECTOR, "h6"),
    ]
    for by, value in candidates:
        try:
            el = driver.find_element(by, value)
            text = (el.text or "").strip()
            if text:
                return "css", value, text
        except Exception:
            continue
    return "", "", ""


def _to_priority(element_type: str) -> str:
    et = (element_type or "").lower()
    if "submit" in et or "button" in et:
        return "High"
    if "link" in et or et in ("a", "span"):
        return "Medium"
    if et in ("text", "password", "email", "input"):
        return "High"
    return "Low"


def _step(
    tc_id: str,
    tc_name: str,
    step_no: int,
    action: str,
    locator_type: str = "",
    locator_value: str = "",
    test_data: str = "",
    expected_result: str = "",
    run: str = "Yes",
    priority: str = "Medium",
    element_type: str = "",
) -> List[str]:
    return [
        tc_id,
        tc_name,
        step_no,
        action,
        locator_type,
        locator_value,
        test_data,
        expected_result,
        run,
        priority,
        element_type,
        "Not Executed",
        "",
        "",
    ]


def _write_excel(output_path: Path, rows: List[List[str]], target_url: str) -> None:
    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = "Test Plan"
    ws_plan.column_dimensions["A"].width = 22
    ws_plan.column_dimensions["B"].width = 55
    ws_plan.column_dimensions["C"].width = 22
    ws_plan.column_dimensions["D"].width = 35

    title_font = Font(bold=True, size=16, color="1F4E79")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    label_fill = PatternFill(start_color="EAF2FA", end_color="EAF2FA", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws_plan.merge_cells("A1:D1")
    ws_plan["A1"] = "AUTO-GENERATED TEST PLAN"
    ws_plan["A1"].font = title_font
    ws_plan["A1"].alignment = Alignment(horizontal="center")

    ws_plan.merge_cells("A3:D3")
    ws_plan["A3"] = "PROJECT INFORMATION"
    ws_plan["A3"].font = section_font
    ws_plan["A3"].fill = section_fill
    ws_plan["A3"].alignment = Alignment(horizontal="left")
    for col in "ABCD":
        ws_plan[f"{col}3"].border = thin

    details = [
        ("Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Target URL", target_url),
        ("Generated By", "Auto Script Generator", "Suite Type", "Keyword-Driven Selenium"),
        ("Sheets", "Test Plan, Test Cases", "Execution", "Update 'Run' column then execute"),
    ]
    start_row = 4
    for idx, (k1, v1, k2, v2) in enumerate(details):
        r = start_row + idx
        ws_plan.cell(r, 1, k1).fill = label_fill
        ws_plan.cell(r, 1).border = thin
        ws_plan.cell(r, 1).font = Font(bold=True)
        ws_plan.cell(r, 2, v1).border = thin
        ws_plan.cell(r, 3, k2).fill = label_fill
        ws_plan.cell(r, 3).border = thin
        ws_plan.cell(r, 3).font = Font(bold=True)
        ws_plan.cell(r, 4, v2).border = thin

    ws = wb.create_sheet("Test Cases")
    headers = BASE_HEADERS + EXTRA_HEADERS
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    alt_fill = PatternFill(start_color="F7FAFD", end_color="F7FAFD", fill_type="solid")
    for r, row in enumerate(rows, 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(r, c, value)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            if r % 2 == 0:
                cell.fill = alt_fill

    col_widths = [14, 28, 9, 22, 14, 42, 26, 26, 8, 11, 14, 14, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def generate_excel_from_site(
    base_url: str,
    username: str,
    password: str,
    browser: str,
    output_path: Path,
    max_pages: int = 8,
    generation_mode: str = "auto_login_navigation",
    logger: Optional[Callable[[str], None]] = None,
) -> Dict[str, str]:
    if not base_url:
        raise ValueError("Base URL is required")

    driver = None
    generated_rows: List[List[str]] = []
    try:
        mode = (generation_mode or "auto_login_navigation").strip().lower()
        if mode not in {"auto_login_navigation", "navigation_only", "minimal_smoke"}:
            mode = "auto_login_navigation"

        _log(logger, f"Launching {browser} browser...")
        driver = _create_driver(browser)
        driver.maximize_window()
        _log(logger, f"Opening target URL: {base_url}")
        driver.get(base_url)
        time.sleep(2)

        if mode == "minimal_smoke":
            _log(logger, "Generation mode: minimal_smoke")
            generated_rows = _build_minimal_smoke_rows()
            _write_excel(output_path, generated_rows, base_url)
            _log(logger, f"Excel test script generated: {output_path} (2 test cases, {len(generated_rows)} steps)")
            return {"output_path": str(output_path), "total_steps": str(len(generated_rows))}

        username_el, u_by, u_val = _find_first_visible(
            driver,
            [
                (By.NAME, "username"),
                (By.ID, "username"),
                (By.CSS_SELECTOR, "input[id*='user' i]"),
                (By.CSS_SELECTOR, "input[name*='user' i]"),
                (By.CSS_SELECTOR, "input[id*='email' i]"),
                (By.CSS_SELECTOR, "input[name*='email' i]"),
                (By.CSS_SELECTOR, "input[type='email']"),
                (By.CSS_SELECTOR, "input[type='text']"),
            ],
        )
        password_el, p_by, p_val = _find_first_visible(
            driver,
            [
                (By.NAME, "password"),
                (By.ID, "password"),
                (By.CSS_SELECTOR, "input[id*='pass' i]"),
                (By.CSS_SELECTOR, "input[name*='pass' i]"),
                (By.CSS_SELECTOR, "input[type='password']"),
            ],
        )
        submit_el, s_by, s_val = _find_first_visible(
            driver,
            [
                (By.CSS_SELECTOR, "button[type='submit']"),
                (By.CSS_SELECTOR, "input[type='submit']"),
                (By.CSS_SELECTOR, "button[id*='login' i]"),
                (By.CSS_SELECTOR, "button[class*='login' i]"),
                (By.CSS_SELECTOR, "button[onclick*='login' i]"),
                (By.XPATH, "//button[contains(.,'Login') or contains(.,'Sign in')]"),
            ],
        )

        # Some websites (e.g., modal login) need clicking a login trigger first.
        if not username_el or not password_el or not submit_el:
            trigger, _, _ = _find_first_visible(
                driver,
                [
                    (By.CSS_SELECTOR, "a[id*='login' i]"),
                    (By.CSS_SELECTOR, "button[id*='login' i]"),
                    (By.CSS_SELECTOR, "a[href*='login' i]"),
                    (By.CSS_SELECTOR, "[data-target*='login' i]"),
                    (By.CSS_SELECTOR, "[onclick*='login' i]"),
                    (By.XPATH, "//a[contains(., 'Log in') or contains(., 'Login')]"),
                    (By.XPATH, "//button[contains(., 'Log in') or contains(., 'Login')]"),
                ],
                timeout=5,
            )
            if trigger:
                try:
                    trigger.click()
                    time.sleep(1.5)
                except Exception:
                    pass

            username_el, u_by, u_val = _find_first_visible(
                driver,
                [
                    (By.NAME, "username"),
                    (By.ID, "username"),
                    (By.CSS_SELECTOR, "input[id*='user' i]"),
                    (By.CSS_SELECTOR, "input[name*='user' i]"),
                    (By.CSS_SELECTOR, "input[id*='email' i]"),
                    (By.CSS_SELECTOR, "input[name*='email' i]"),
                    (By.CSS_SELECTOR, "input[type='email']"),
                    (By.CSS_SELECTOR, "input[type='text']"),
                ],
                timeout=6,
            )
            password_el, p_by, p_val = _find_first_visible(
                driver,
                [
                    (By.NAME, "password"),
                    (By.ID, "password"),
                    (By.CSS_SELECTOR, "input[id*='pass' i]"),
                    (By.CSS_SELECTOR, "input[name*='pass' i]"),
                    (By.CSS_SELECTOR, "input[type='password']"),
                ],
                timeout=6,
            )
            submit_el, s_by, s_val = _find_first_visible(
                driver,
                [
                    (By.CSS_SELECTOR, "button[type='submit']"),
                    (By.CSS_SELECTOR, "input[type='submit']"),
                    (By.CSS_SELECTOR, "button[id*='login' i]"),
                    (By.CSS_SELECTOR, "button[class*='login' i]"),
                    (By.CSS_SELECTOR, "button[onclick*='login' i]"),
                    (By.XPATH, "//button[contains(.,'Login') or contains(.,'Sign in') or contains(.,'Log in')]"),
                ],
                timeout=6,
            )

        login_available = bool(username_el and password_el and submit_el) and mode != "navigation_only"
        login_locators = None
        if login_available:
            login_locators = {
                "username": (u_by, u_val),
                "password": (p_by, p_val),
                "submit": (s_by, s_val),
            }
            _log(logger, "Generation mode: auto_login_navigation")
            _log(logger, "Login form detected. Building login-based test cases...")
            login_tc = "TC001"
            generated_rows.extend(
                [
                    _step(login_tc, "Valid Login", 1, "navigate", test_data="${base_url}", priority="High"),
                    _step(login_tc, "Valid Login", 2, "type", _by_to_locator_type(u_by), u_val, "${username}", priority="High", element_type="text"),
                    _step(login_tc, "Valid Login", 3, "type", _by_to_locator_type(p_by), p_val, "${password}", priority="High", element_type="password"),
                    _step(login_tc, "Valid Login", 4, "click", _by_to_locator_type(s_by), s_val, priority="High", element_type="button"),
                    _step(login_tc, "Valid Login", 5, "wait", test_data="3", priority="Medium"),
                    _step(login_tc, "Valid Login", 6, "verify_url", expected_result="/", priority="Medium"),
                    _step(login_tc, "Valid Login", 7, "take_screenshot", test_data="TC001_valid_login", priority="Low"),
                ]
            )

        if login_available:
            username_el.clear()
            username_el.send_keys(username)
            password_el.clear()
            password_el.send_keys(password)
            submit_el.click()
            time.sleep(4)
            post_login_url = driver.current_url
            _log(logger, f"Login attempted. Current URL: {post_login_url}")
        else:
            if mode == "navigation_only":
                _log(logger, "Generation mode: navigation_only")
                _log(logger, "Login is intentionally skipped in navigation_only mode.")
            else:
                _log(logger, "Login form not detected. Falling back to public-page script generation.")

        nav_candidates = _discover_nav_candidates(driver, max_pages=max_pages)
        if not nav_candidates:
            _log(logger, "No nav candidates found after first scan, retrying from home page...")
            driver.get(base_url)
            time.sleep(2)
            nav_candidates = _discover_nav_candidates(driver, max_pages=max_pages)

        _log(logger, f"Discovered {len(nav_candidates)} navigation candidates.")

        tc_start = 2
        generated_cases = 1 if login_available else 0
        for idx, nav in enumerate(nav_candidates, tc_start):
            tc_id = f"TC{idx:03d}"
            tc_name = f"Navigate to {nav['text']}"
            _log(logger, f"Profiling page action: {nav['text']}")

            # Re-login quickly to keep each generated test isolated.
            if login_available:
                logged_in = _safe_login(driver, base_url, username, password, login_locators)
                if not logged_in:
                    _log(logger, f"Login refresh failed for {nav['text']}; generating generic step set.")
                    generated_rows.extend(
                        _rowset_for_nav(
                            tc_id=tc_id,
                            tc_name=tc_name,
                            login_locators=login_locators,
                            nav=nav,
                            nav_url_fragment="",
                            heading_selector="",
                            heading_text="",
                        )
                    )
                    generated_cases += 1
                    continue
            else:
                driver.get(base_url)
                time.sleep(2)

            nav_url_fragment = ""
            heading_selector = ""
            heading_text = ""
            try:
                click_el = driver.find_element(nav["locator_type"], nav["locator_value"])
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", click_el)
                time.sleep(0.5)
                click_el.click()
                time.sleep(2)
                current = driver.current_url
                nav_url_fragment = current.rstrip("/").split("/")[-1] if current else ""
                heading_type, heading_selector, heading_text = _extract_heading(driver)
            except Exception:
                pass

            rows = _rowset_for_nav(
                tc_id=tc_id,
                tc_name=tc_name,
                login_locators=login_locators if login_available else None,
                nav=nav,
                nav_url_fragment=nav_url_fragment,
                heading_selector=heading_selector,
                heading_text=heading_text,
            )
            generated_rows.extend(rows)
            generated_cases += 1

        if generated_cases == 0:
            # Minimal fallback so generation never fails for new URLs.
            generated_rows.extend(
                [
                    _step("TC001", "Basic URL Check", 1, "navigate", test_data="${base_url}", priority="High"),
                    _step("TC001", "Basic URL Check", 2, "wait", test_data="2", priority="Medium"),
                    _step("TC001", "Basic URL Check", 3, "verify_url", expected_result="/", priority="Medium"),
                    _step("TC001", "Basic URL Check", 4, "take_screenshot", test_data="TC001_basic_url_check", priority="Low"),
                ]
            )
            generated_cases = 1
            _log(logger, "Applied minimal fallback test case because no navigation elements were discoverable.")

        _write_excel(output_path, generated_rows, base_url)
        _log(logger, f"Excel test script generated: {output_path} ({generated_cases} test cases, {len(generated_rows)} steps)")
        return {"output_path": str(output_path), "total_steps": str(len(generated_rows))}
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
