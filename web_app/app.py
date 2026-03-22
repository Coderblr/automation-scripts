import os
import sys
import json
import shutil
import threading
import subprocess
import time
from pathlib import Path
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
try:
    from auto_script_generator import generate_excel_from_site
except ModuleNotFoundError:
    from web_app.auto_script_generator import generate_excel_from_site

PROJECT_ROOT = Path(__file__).parent.parent.resolve()
CONFIG_PATH = PROJECT_ROOT / "config" / "config.ini"
TEST_DATA_DIR = PROJECT_ROOT / "test_data"
REPORTS_DIR = PROJECT_ROOT / "reports"
PYTHON_FW_DIR = PROJECT_ROOT / "python_framework"
JAVA_FW_DIR = PROJECT_ROOT / "java_framework"
STATIC_DIR = Path(__file__).parent / "static"

app = FastAPI(title="Test Automation Dashboard")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# --------------- Global runner state ---------------
runner = {
    "status": "idle",
    "lines": [],
    "method": "",
    "browser": "",
    "start_time": None,
    "end_time": None,
    "exit_code": None,
    "process": None,
}
lock = threading.Lock()

generator_state = {
    "status": "idle",
    "lines": [],
    "start_time": None,
    "end_time": None,
    "output_file": "",
    "error": "",
}


# --------------- Models ---------------
class ConfigUpdate(BaseModel):
    base_url: str = ""
    api_base_url: str = ""
    test_username: str = ""
    test_password: str = ""
    browser_name: str = "chrome"
    headless: bool = False


class RunRequest(BaseModel):
    method: str = "python"
    browser: str = "chrome"


class GenerateScriptsRequest(BaseModel):
    base_url: str
    username: str
    password: str
    browser: str = "chrome"
    max_pages: int = 12
    generation_mode: str = "auto_login_navigation"


# --------------- Config helpers ---------------
def _parse_ini(path: Path) -> dict:
    data = {}
    if not path.exists():
        return data
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("["):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                data[k.strip()] = v.strip()
    return data


def _read_config() -> dict:
    d = _parse_ini(CONFIG_PATH)
    return {
        "base_url": d.get("base_url", ""),
        "api_base_url": d.get("api_base_url", "http://localhost:8000"),
        "test_username": d.get("test_username", ""),
        "test_password": d.get("test_password", ""),
        "browser_name": d.get("browser_name", "chrome"),
        "headless": d.get("headless", "false").lower() == "true",
    }


def _write_config(d: dict):
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    content = (
        "[browser]\n"
        "# Supported: chrome, firefox, edge\n"
        f"browser_name = {d.get('browser_name', 'chrome')}\n"
        f"headless = {str(d.get('headless', False)).lower()}\n"
        "implicit_wait = 10\n"
        "explicit_wait = 20\n"
        "page_load_timeout = 30\n"
        "\n"
        "[application]\n"
        f"base_url = {d.get('base_url', '')}\n"
        f"api_base_url = {d.get('api_base_url', 'http://localhost:8000')}\n"
        "\n"
        "[test_data]\n"
        "excel_path = test_data/test_cases.xlsx\n"
        "screenshot_on_failure = true\n"
        "\n"
        "[reporting]\n"
        "report_path = reports\n"
        "generate_html_report = true\n"
        "\n"
        "[api]\n"
        "auth_endpoint = /api/auth/login\n"
        "health_endpoint = /api/health\n"
        "\n"
        "[credentials]\n"
        f"test_username = {d.get('test_username', '')}\n"
        f"test_password = {d.get('test_password', '')}\n"
    )
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        f.write(content)


# --------------- Background runner ---------------
def _run_in_background(method: str, browser: str):
    with lock:
        runner["status"] = "running"
        runner["lines"] = []
        runner["method"] = method
        runner["browser"] = browser
        runner["start_time"] = datetime.now().isoformat()
        runner["end_time"] = None
        runner["exit_code"] = None

    try:
        if method == "python":
            cmd = [
                sys.executable, "-m", "pytest",
                "tests/test_excel_driven.py::TestExcelDriven::test_run_excel_test_cases",
                "-v", "--tb=short", "--color=no",
                "--html=../reports/test_report.html", "--self-contained-html",
            ]
            cwd = str(PYTHON_FW_DIR)
            use_shell = False
        else:
            cmd = f'mvn clean test -Dbrowser={browser}'
            cwd = str(JAVA_FW_DIR)
            use_shell = True

        env = os.environ.copy()
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, cwd=cwd, bufsize=1, shell=use_shell, env=env,
        )

        with lock:
            runner["process"] = proc

        for line in iter(proc.stdout.readline, ""):
            with lock:
                runner["lines"].append(line.rstrip("\n\r"))

        proc.wait()

        with lock:
            runner["status"] = "completed" if proc.returncode == 0 else "failed"
            runner["exit_code"] = proc.returncode
            runner["end_time"] = datetime.now().isoformat()
            runner["process"] = None

    except Exception as exc:
        with lock:
            runner["status"] = "failed"
            runner["lines"].append(f"ERROR: {exc}")
            runner["end_time"] = datetime.now().isoformat()
            runner["process"] = None


def _generate_scripts_background(req: GenerateScriptsRequest):
    with lock:
        generator_state["status"] = "running"
        generator_state["lines"] = []
        generator_state["start_time"] = datetime.now().isoformat()
        generator_state["end_time"] = None
        generator_state["output_file"] = ""
        generator_state["error"] = ""

    def gen_log(line: str):
        with lock:
            generator_state["lines"].append(line)

    try:
        TEST_DATA_DIR.mkdir(parents=True, exist_ok=True)
        output_file = TEST_DATA_DIR / "test_cases.xlsx"

        # Keep old generated file as backup if it exists.
        if output_file.exists():
            backup_name = f"test_cases_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_file.replace(TEST_DATA_DIR / backup_name)
            gen_log(f"Backed up existing Excel to {backup_name}")

        result = generate_excel_from_site(
            base_url=req.base_url,
            username=req.username,
            password=req.password,
            browser=req.browser,
            output_path=output_file,
            max_pages=max(2, min(req.max_pages, 30)),
            generation_mode=req.generation_mode,
            logger=gen_log,
        )

        with lock:
            generator_state["status"] = "completed"
            generator_state["end_time"] = datetime.now().isoformat()
            generator_state["output_file"] = result["output_path"]
    except Exception as exc:
        with lock:
            generator_state["status"] = "failed"
            generator_state["end_time"] = datetime.now().isoformat()
            generator_state["error"] = str(exc)
            generator_state["lines"].append(f"ERROR: {exc}")


# --------------- API endpoints ---------------
@app.get("/api/config")
def get_config():
    return _read_config()


@app.post("/api/config")
def update_config(cfg: ConfigUpdate):
    _write_config(cfg.dict())
    return {"message": "Configuration saved", "config": _read_config()}


@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx / .xls files are accepted")
    TEST_DATA_DIR.mkdir(parents=True, exist_ok=True)
    dest = TEST_DATA_DIR / "test_cases.xlsx"
    with open(dest, "wb") as f:
        f.write(await file.read())
    return {"message": f"Uploaded {file.filename}", "path": str(dest)}


@app.post("/api/generate-excel-scripts")
def generate_excel_scripts(req: GenerateScriptsRequest):
    with lock:
        if generator_state["status"] == "running":
            raise HTTPException(409, "Excel generation already running")

    cfg = _read_config()
    cfg["base_url"] = req.base_url.strip()
    cfg["test_username"] = req.username.strip()
    cfg["test_password"] = req.password.strip()
    cfg["browser_name"] = req.browser.strip().lower() or "chrome"
    _write_config(cfg)

    t = threading.Thread(target=_generate_scripts_background, args=(req,), daemon=True)
    t.start()
    return {"message": "Excel generation started"}


@app.get("/api/generate-excel-status")
def generate_excel_status():
    with lock:
        return {
            "status": generator_state["status"],
            "start_time": generator_state["start_time"],
            "end_time": generator_state["end_time"],
            "output_file": os.path.basename(generator_state["output_file"]) if generator_state["output_file"] else "",
            "error": generator_state["error"],
            "logs": generator_state["lines"][-400:],
        }


@app.get("/api/generated-excel/download")
def download_generated_excel():
    with lock:
        output = generator_state["output_file"]
    if not output:
        raise HTTPException(404, "No generated Excel found")
    fp = Path(output)
    if not fp.exists():
        raise HTTPException(404, "Generated Excel file is missing")
    return FileResponse(str(fp), filename=fp.name)


@app.post("/api/run-tests")
def run_tests(req: RunRequest):
    with lock:
        if runner["status"] == "running":
            raise HTTPException(409, "Tests are already running")

    cfg = _read_config()
    cfg["browser_name"] = req.browser
    _write_config(cfg)

    t = threading.Thread(target=_run_in_background, args=(req.method, req.browser), daemon=True)
    t.start()
    return {"message": "Test execution started", "method": req.method, "browser": req.browser}


@app.get("/api/run-status")
def run_status():
    with lock:
        return {
            "status": runner["status"],
            "method": runner["method"],
            "browser": runner["browser"],
            "start_time": runner["start_time"],
            "end_time": runner["end_time"],
            "exit_code": runner["exit_code"],
            "output": runner["lines"][-500:],
            "total_lines": len(runner["lines"]),
        }


@app.post("/api/stop-tests")
def stop_tests():
    with lock:
        if runner["process"] is not None:
            runner["process"].terminate()
            runner["status"] = "stopped"
            runner["end_time"] = datetime.now().isoformat()
            return {"message": "Tests stopped"}
    raise HTTPException(400, "No tests are currently running")


@app.get("/api/reports")
def list_reports():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    reports = []

    for pattern in ("*.html", "*.xlsx"):
        for fp in REPORTS_DIR.glob(pattern):
            reports.append({
                "name": fp.name,
                "size": fp.stat().st_size,
                "modified": datetime.fromtimestamp(fp.stat().st_mtime).isoformat(),
                "type": "html" if fp.suffix == ".html" else "excel",
            })

    testng_dir = JAVA_FW_DIR / "target" / "surefire-reports"
    if testng_dir.exists():
        for fp in testng_dir.glob("*.html"):
            reports.append({
                "name": f"java/{fp.name}",
                "size": fp.stat().st_size,
                "modified": datetime.fromtimestamp(fp.stat().st_mtime).isoformat(),
                "type": "html",
            })

    reports.sort(key=lambda r: r["modified"], reverse=True)
    return reports


@app.get("/api/reports/{report_name:path}")
def get_report(report_name: str):
    if report_name.startswith("java/"):
        fp = JAVA_FW_DIR / "target" / "surefire-reports" / report_name[5:]
    else:
        fp = REPORTS_DIR / report_name
    if not fp.exists():
        raise HTTPException(404, "Report not found")
    return FileResponse(str(fp))


@app.get("/api/test-cases")
def get_test_cases():
    excel_path = TEST_DATA_DIR / "test_cases.xlsx"
    if not excel_path.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path), read_only=True)
        ws = wb["Test Cases"] if "Test Cases" in wb.sheetnames else wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        cases = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rd = dict(zip(headers, row))
            tc_id = str(rd.get("Test_Case_ID") or "")
            tc_name = str(rd.get("Test_Case_Name") or "")
            if tc_id:
                if tc_id not in cases:
                    cases[tc_id] = {"id": tc_id, "name": tc_name, "steps": 0}
                cases[tc_id]["steps"] += 1
        wb.close()
        return list(cases.values())
    except Exception as exc:
        return [{"id": "error", "name": str(exc), "steps": 0}]


@app.get("/api/check-prerequisites")
def check_prerequisites():
    checks = {}

    try:
        r = subprocess.run([sys.executable, "--version"], capture_output=True, text=True, timeout=10)
        checks["python"] = {"installed": True, "version": r.stdout.strip() or r.stderr.strip()}
    except Exception:
        checks["python"] = {"installed": False, "version": ""}

    try:
        r = subprocess.run(["java", "-version"], capture_output=True, text=True, timeout=10)
        ver = (r.stderr or r.stdout).strip().split("\n")[0]
        checks["java"] = {"installed": True, "version": ver}
    except Exception:
        checks["java"] = {"installed": False, "version": ""}

    try:
        r = subprocess.run("mvn -version", capture_output=True, text=True, timeout=10, shell=True)
        ver = (r.stdout or r.stderr).strip().split("\n")[0]
        checks["maven"] = {"installed": True, "version": ver}
    except Exception:
        checks["maven"] = {"installed": False, "version": ""}

    checks["excel"] = {"exists": (TEST_DATA_DIR / "test_cases.xlsx").exists()}
    return checks


# --------------- Serve static frontend ---------------
app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    print("\n  Test Automation Dashboard")
    print("  http://localhost:5000\n")
    uvicorn.run(app, host="0.0.0.0", port=5000)
