"""
Excel Handler for reading test cases and writing results.

Expected Excel format (columns):
  Test_Case_ID | Test_Case_Name | Step_No | Action | Locator_Type | Locator_Value | Test_Data | Expected_Result | Run

Run column: 'Yes' to run, 'No' to skip
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = [
    'Test_Case_ID', 'Test_Case_Name', 'Step_No', 'Action',
    'Locator_Type', 'Locator_Value', 'Test_Data', 'Expected_Result', 'Run'
]


class ExcelHandler:
    """Reads test cases from Excel and writes test results back."""

    def __init__(self, file_path):
        self.file_path = os.path.abspath(file_path)
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

    def read_test_cases(self, sheet_name=None):
        """
        Read all test cases from the Excel file.
        Returns a list of dicts, each representing a test step.
        Defaults to the 'Test Cases' sheet if available.
        """
        wb = load_workbook(self.file_path, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        elif 'Test Cases' in wb.sheetnames:
            ws = wb['Test Cases']
        else:
            ws = wb.active

        headers = [cell.value for cell in ws[1]]
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ValueError(
                f"Missing required columns: {missing}. "
                f"Required: {REQUIRED_COLUMNS}. Found: {headers}"
            )

        test_steps = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data = {headers[i]: row[i].value for i in range(len(headers))}

            if not row_data.get('Action'):
                continue

            run_flag = str(row_data.get('Run', 'Yes')).strip().lower()
            if run_flag in ('no', 'n', 'false', '0', 'skip'):
                continue

            for key in ['Locator_Type', 'Locator_Value', 'Test_Data', 'Expected_Result']:
                if row_data.get(key) is None:
                    row_data[key] = ''

            test_steps.append(row_data)

        wb.close()
        logger.info(f"Loaded {len(test_steps)} test steps from '{self.file_path}'")
        return test_steps

    def get_test_cases_grouped(self, sheet_name=None):
        """
        Group test steps by Test_Case_ID.
        Returns: dict of {test_case_id: {'name': str, 'steps': [list of step dicts]}}
        """
        all_steps = self.read_test_cases(sheet_name)
        grouped = {}
        for step in all_steps:
            tc_id = step['Test_Case_ID']
            if tc_id not in grouped:
                grouped[tc_id] = {
                    'name': step['Test_Case_Name'],
                    'steps': []
                }
            grouped[tc_id]['steps'].append(step)

        logger.info(f"Found {len(grouped)} test cases")
        return grouped

    def get_sheet_names(self):
        wb = load_workbook(self.file_path, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    @staticmethod
    def write_results(results, output_path=None):
        """
        Write test execution results to an Excel file.
        results: list of dicts with keys: Test_Case_ID, Test_Case_Name, Step_No,
                 Action, Status, Error_Message, Timestamp, Duration
        """
        if not output_path:
            output_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports')
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(output_dir, f'test_results_{timestamp}.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Results'

        headers = ['Test_Case_ID', 'Test_Case_Name', 'Step_No', 'Action',
                    'Status', 'Error_Message', 'Timestamp', 'Duration_ms']

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        skip_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, result in enumerate(results, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=result.get(header, ''))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            status = str(result.get('Status', '')).upper()
            status_cell = ws.cell(row=row_idx, column=headers.index('Status') + 1)
            if status == 'PASS':
                status_cell.fill = pass_fill
            elif status == 'FAIL':
                status_cell.fill = fail_fill
            elif status == 'SKIP':
                status_cell.fill = skip_fill

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        ws.auto_filter.ref = ws.dimensions
        wb.save(output_path)
        logger.info(f"Results written to: {output_path}")
        return output_path


def _build_test_plan_sheet(wb):
    """
    Create a professional Test Plan cover sheet matching standard QA documentation format.
    Includes: Title, Description, Prepared By, Date, Tools Used, Environment, Scope, etc.
    """
    ws = wb.active
    ws.title = 'Test Plan'

    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=16, color='1F4E79')
    section_font = Font(bold=True, size=12, color='FFFFFF')
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    section_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35

    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value='AUTOMATION TEST PLAN')
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:D2')
    subtitle = ws.cell(row=2, column=1, value='SauceDemo Web Application - Regression Test Suite')
    subtitle.font = Font(size=12, italic=True, color='4472C4')
    subtitle.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25

    # Section: Project Information
    row = 4
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value='PROJECT INFORMATION' if col == 1 else '')
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
    ws.merge_cells(f'A{row}:D{row}')

    info_data = [
        ('Title', 'SauceDemo Automation Test Suite', 'Project ID', 'SDEMO-ATS-001'),
        ('Project Description', 'This document contains automated regression test scripts for the SauceDemo e-commerce web application. It covers login, product browsing, cart management, checkout flow, sorting, and logout functionalities.', '', ''),
        ('Prepared By', 'QA Team', 'Date', datetime.now().strftime('%Y-%m-%d')),
        ('Tools Used', 'Python Selenium / Java Selenium (Automated)', 'Version', '1.0'),
        ('Environment', 'Staging (https://www.saucedemo.com/)', 'Browser', 'Chrome / Firefox / Edge'),
        ('Test Type', 'Regression, Smoke, Functional, E2E', 'Priority', 'High'),
    ]

    for i, (label1, value1, label2, value2) in enumerate(info_data):
        r = row + 1 + i
        c1 = ws.cell(row=r, column=1, value=label1)
        c1.font = label_font
        c1.fill = light_fill
        c1.border = thin_border
        c1.alignment = Alignment(wrap_text=True, vertical='top')

        c2 = ws.cell(row=r, column=2, value=value1)
        c2.font = value_font
        c2.border = thin_border
        c2.alignment = Alignment(wrap_text=True, vertical='top')

        if label2:
            c3 = ws.cell(row=r, column=3, value=label2)
            c3.font = label_font
            c3.fill = light_fill
            c3.border = thin_border
            c3.alignment = Alignment(wrap_text=True, vertical='top')

            c4 = ws.cell(row=r, column=4, value=value2)
            c4.font = value_font
            c4.border = thin_border
            c4.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            ws.merge_cells(f'B{r}:D{r}')
            c2.alignment = Alignment(wrap_text=True, vertical='top')

        ws.row_dimensions[r].height = 30 if label1 != 'Project Description' else 50

    # Section: Test Scope
    row = 12
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value='TEST SCOPE' if col == 1 else '')
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
    ws.merge_cells(f'A{row}:D{row}')

    scope_items = [
        ('TC001', 'Valid Login', 'Verify successful login with valid credentials', 'Regression'),
        ('TC002', 'Invalid Login', 'Verify error message for invalid credentials', 'Regression'),
        ('TC003', 'Locked Out User', 'Verify locked out user gets proper error', 'Regression'),
        ('TC004', 'Product Page Verification', 'Verify products display on inventory page', 'Smoke'),
        ('TC005', 'Add Product to Cart', 'Verify adding a product to shopping cart', 'Functional'),
        ('TC006', 'Remove Product from Cart', 'Verify removing a product from cart', 'Functional'),
        ('TC007', 'Complete Checkout Flow', 'Verify full checkout process end-to-end', 'E2E'),
        ('TC008', 'Product Sorting', 'Verify sorting products by name and price', 'Functional'),
        ('TC009', 'Logout', 'Verify user can logout from the application', 'Regression'),
        ('TC010', 'Cart Badge Count', 'Verify cart badge updates correctly', 'Functional'),
        ('TC011', 'Product Details Navigation', 'Verify navigating to product detail page', 'Functional'),
        ('TC012', 'Empty Cart Checkout', 'Verify checkout behavior with empty cart', 'Negative'),
    ]

    scope_headers = ['Test Case ID', 'Test Case Name', 'Description', 'Test Type']
    for col, header in enumerate(scope_headers, 1):
        cell = ws.cell(row=row + 1, column=col, value=header)
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for i, (tc_id, name, desc, ttype) in enumerate(scope_items):
        r = row + 2 + i
        for col, val in enumerate([tc_id, name, desc, ttype], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if i % 2 == 0:
                cell.fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    # Section: Test Credentials
    cred_row = row + 2 + len(scope_items) + 1
    for col in range(1, 5):
        cell = ws.cell(row=cred_row, column=col, value='TEST CREDENTIALS' if col == 1 else '')
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
    ws.merge_cells(f'A{cred_row}:D{cred_row}')

    creds = [
        ('standard_user', 'secret_sauce', 'Standard user - full access', 'Active'),
        ('locked_out_user', 'secret_sauce', 'Locked out user - login denied', 'Locked'),
        ('problem_user', 'secret_sauce', 'Problem user - buggy behavior', 'Active'),
        ('performance_glitch_user', 'secret_sauce', 'Slow performance user', 'Active'),
    ]

    cred_headers = ['Username', 'Password', 'Description', 'Status']
    for col, header in enumerate(cred_headers, 1):
        cell = ws.cell(row=cred_row + 1, column=col, value=header)
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for i, (user, pw, desc, status) in enumerate(creds):
        r = cred_row + 2 + i
        for col, val in enumerate([user, pw, desc, status], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    return ws


def create_test_case_template(output_path=None):
    """Create a comprehensive Excel template with Test Plan + SauceDemo test cases."""
    if not output_path:
        output_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'test_data')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, 'test_cases.xlsx')

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ═══════════════════════════════════════════
    # Sheet 1: Test Plan Cover Sheet
    # ═══════════════════════════════════════════
    _build_test_plan_sheet(wb)

    # ═══════════════════════════════════════════
    # Sheet 2: Test Cases (SauceDemo)
    # ═══════════════════════════════════════════
    ws = wb.create_sheet('Test Cases')

    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    headers = REQUIRED_COLUMNS
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Using ${variable} placeholders - values come from config/config.ini
    # Testers only need to change config.ini to test a different application
    BASE = '${base_url}'
    USER = '${username}'
    PASS = '${password}'

    sample_data = [
        # ──── TC001: Valid Login ────
        ['TC001', 'Valid Login', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC001', 'Valid Login', 2, 'verify_title', '', '', 'Swag Labs', '', 'Yes'],
        ['TC001', 'Valid Login', 3, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC001', 'Valid Login', 4, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC001', 'Valid Login', 5, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC001', 'Valid Login', 6, 'wait', '', '', '2', '', 'Yes'],
        ['TC001', 'Valid Login', 7, 'verify_url', '', '', '', 'inventory.html', 'Yes'],
        ['TC001', 'Valid Login', 8, 'verify_element_visible', 'class', 'inventory_list', '', '', 'Yes'],
        ['TC001', 'Valid Login', 9, 'take_screenshot', '', '', 'TC001_login_success', '', 'Yes'],

        # ──── TC002: Invalid Login ────
        ['TC002', 'Invalid Login', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC002', 'Invalid Login', 2, 'type', 'id', 'user-name', 'invalid_user_xyz', '', 'Yes'],
        ['TC002', 'Invalid Login', 3, 'type', 'id', 'password', 'wrong_password_xyz', '', 'Yes'],
        ['TC002', 'Invalid Login', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC002', 'Invalid Login', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC002', 'Invalid Login', 6, 'verify_element_visible', 'css', 'h3[data-test="error"]', '', '', 'Yes'],
        ['TC002', 'Invalid Login', 7, 'verify_text', 'css', 'h3[data-test="error"]', '', 'do not match', 'Yes'],
        ['TC002', 'Invalid Login', 8, 'take_screenshot', '', '', 'TC002_invalid_login_error', '', 'Yes'],

        # ──── TC003: Locked Out User ────
        ['TC003', 'Locked Out User', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC003', 'Locked Out User', 2, 'type', 'id', 'user-name', 'locked_out_user', '', 'Yes'],
        ['TC003', 'Locked Out User', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC003', 'Locked Out User', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC003', 'Locked Out User', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC003', 'Locked Out User', 6, 'verify_element_visible', 'css', 'h3[data-test="error"]', '', '', 'Yes'],
        ['TC003', 'Locked Out User', 7, 'verify_text', 'css', 'h3[data-test="error"]', '', 'locked out', 'Yes'],
        ['TC003', 'Locked Out User', 8, 'take_screenshot', '', '', 'TC003_locked_out_error', '', 'Yes'],

        # ──── TC004: Product Page Verification ────
        ['TC004', 'Product Page Verification', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC004', 'Product Page Verification', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC004', 'Product Page Verification', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC004', 'Product Page Verification', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC004', 'Product Page Verification', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC004', 'Product Page Verification', 6, 'verify_text', 'css', '.header_secondary_container .title', '', 'Products', 'Yes'],
        ['TC004', 'Product Page Verification', 7, 'verify_element_visible', 'class', 'inventory_list', '', '', 'Yes'],
        ['TC004', 'Product Page Verification', 8, 'verify_element_visible', 'class', 'product_sort_container', '', '', 'Yes'],
        ['TC004', 'Product Page Verification', 9, 'verify_element_visible', 'css', '.shopping_cart_link', '', '', 'Yes'],
        ['TC004', 'Product Page Verification', 10, 'take_screenshot', '', '', 'TC004_products_page', '', 'Yes'],

        # ──── TC005: Add Product to Cart ────
        ['TC005', 'Add Product to Cart', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC005', 'Add Product to Cart', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC005', 'Add Product to Cart', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC005', 'Add Product to Cart', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 6, 'click', 'css', '.inventory_item:nth-child(1) .btn_inventory', '', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 7, 'wait', '', '', '2', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 8, 'verify_element_visible', 'css', '.shopping_cart_badge', '', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 9, 'verify_text', 'css', '.shopping_cart_badge', '', '1', 'Yes'],
        ['TC005', 'Add Product to Cart', 10, 'click', 'css', '.shopping_cart_link', '', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 11, 'wait', '', '', '2', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 12, 'verify_url', '', '', '', 'cart.html', 'Yes'],
        ['TC005', 'Add Product to Cart', 13, 'verify_element_visible', 'css', '.cart_item', '', '', 'Yes'],
        ['TC005', 'Add Product to Cart', 14, 'take_screenshot', '', '', 'TC005_product_in_cart', '', 'Yes'],

        # ──── TC006: Remove Product from Cart ────
        ['TC006', 'Remove Product from Cart', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 6, 'click', 'css', '.inventory_item:nth-child(1) .btn_inventory', '', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 7, 'wait', '', '', '2', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 8, 'verify_text', 'css', '.shopping_cart_badge', '', '1', 'Yes'],
        ['TC006', 'Remove Product from Cart', 9, 'click', 'css', '.shopping_cart_link', '', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 10, 'wait', '', '', '2', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 11, 'click', 'css', '.cart_item .btn_secondary', '', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 12, 'wait', '', '', '1', '', 'Yes'],
        ['TC006', 'Remove Product from Cart', 13, 'take_screenshot', '', '', 'TC006_cart_after_remove', '', 'Yes'],

        # ──── TC007: Complete Checkout Flow (E2E) ────
        ['TC007', 'Complete Checkout Flow', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 6, 'click', 'css', '.inventory_item:nth-child(1) .btn_inventory', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 7, 'wait', '', '', '1', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 8, 'click', 'css', '.shopping_cart_link', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 9, 'wait', '', '', '2', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 10, 'verify_url', '', '', '', 'cart.html', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 11, 'verify_element_visible', 'css', '.cart_item', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 12, 'click', 'css', '.checkout_button', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 13, 'wait', '', '', '2', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 14, 'verify_url', '', '', '', 'checkout-step-one.html', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 15, 'type', 'id', 'first-name', 'John', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 16, 'type', 'id', 'last-name', 'Doe', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 17, 'type', 'id', 'postal-code', '12345', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 18, 'click', 'css', '.cart_button', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 19, 'wait', '', '', '2', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 20, 'verify_url', '', '', '', 'checkout-step-two.html', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 21, 'verify_element_visible', 'class', 'summary_info', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 22, 'take_screenshot', '', '', 'TC007_checkout_summary', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 23, 'click', 'css', '.cart_button', '', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 24, 'wait', '', '', '2', '', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 25, 'verify_url', '', '', '', 'checkout-complete.html', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 26, 'verify_text', 'class', 'complete-header', '', 'Thank you for your order', 'Yes'],
        ['TC007', 'Complete Checkout Flow', 27, 'take_screenshot', '', '', 'TC007_order_complete', '', 'Yes'],

        # ──── TC008: Product Sorting ────
        ['TC008', 'Product Sorting', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC008', 'Product Sorting', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC008', 'Product Sorting', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC008', 'Product Sorting', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC008', 'Product Sorting', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC008', 'Product Sorting', 6, 'select_dropdown', 'class', 'product_sort_container', 'Price (low to high)', '', 'Yes'],
        ['TC008', 'Product Sorting', 7, 'wait', '', '', '1', '', 'Yes'],
        ['TC008', 'Product Sorting', 8, 'take_screenshot', '', '', 'TC008_sort_price_low_high', '', 'Yes'],
        ['TC008', 'Product Sorting', 9, 'select_dropdown', 'class', 'product_sort_container', 'Price (high to low)', '', 'Yes'],
        ['TC008', 'Product Sorting', 10, 'wait', '', '', '1', '', 'Yes'],
        ['TC008', 'Product Sorting', 11, 'take_screenshot', '', '', 'TC008_sort_price_high_low', '', 'Yes'],
        ['TC008', 'Product Sorting', 12, 'select_dropdown', 'class', 'product_sort_container', 'Name (Z to A)', '', 'Yes'],
        ['TC008', 'Product Sorting', 13, 'wait', '', '', '1', '', 'Yes'],
        ['TC008', 'Product Sorting', 14, 'take_screenshot', '', '', 'TC008_sort_name_z_to_a', '', 'Yes'],

        # ──── TC009: Logout ────
        ['TC009', 'Logout', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC009', 'Logout', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC009', 'Logout', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC009', 'Logout', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC009', 'Logout', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC009', 'Logout', 6, 'verify_url', '', '', '', 'inventory.html', 'Yes'],
        ['TC009', 'Logout', 7, 'click', 'css', '.bm-burger-button button', '', '', 'Yes'],
        ['TC009', 'Logout', 8, 'wait', '', '', '2', '', 'Yes'],
        ['TC009', 'Logout', 9, 'click', 'id', 'logout_sidebar_link', '', '', 'Yes'],
        ['TC009', 'Logout', 10, 'wait', '', '', '2', '', 'Yes'],
        ['TC009', 'Logout', 11, 'verify_element_visible', 'id', 'login-button', '', '', 'Yes'],
        ['TC009', 'Logout', 12, 'take_screenshot', '', '', 'TC009_logout_success', '', 'Yes'],

        # ──── TC010: Cart Badge Count ────
        ['TC010', 'Cart Badge Count', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC010', 'Cart Badge Count', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC010', 'Cart Badge Count', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC010', 'Cart Badge Count', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 6, 'click', 'css', '.inventory_item:nth-child(1) .btn_inventory', '', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 7, 'wait', '', '', '2', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 8, 'verify_text', 'css', '.shopping_cart_badge', '', '1', 'Yes'],
        ['TC010', 'Cart Badge Count', 9, 'click', 'css', '.inventory_item:nth-child(2) .btn_inventory', '', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 10, 'wait', '', '', '2', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 11, 'verify_text', 'css', '.shopping_cart_badge', '', '2', 'Yes'],
        ['TC010', 'Cart Badge Count', 12, 'click', 'css', '.inventory_item:nth-child(3) .btn_inventory', '', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 13, 'wait', '', '', '2', '', 'Yes'],
        ['TC010', 'Cart Badge Count', 14, 'verify_text', 'css', '.shopping_cart_badge', '', '3', 'Yes'],
        ['TC010', 'Cart Badge Count', 15, 'take_screenshot', '', '', 'TC010_cart_badge_3_items', '', 'Yes'],

        # ──── TC011: Product Details Navigation ────
        ['TC011', 'Product Details Navigation', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC011', 'Product Details Navigation', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC011', 'Product Details Navigation', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC011', 'Product Details Navigation', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 6, 'click', 'css', '.inventory_item:nth-child(1) .inventory_item_name', '', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 7, 'wait', '', '', '2', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 8, 'verify_url', '', '', '', 'inventory-item.html', 'Yes'],
        ['TC011', 'Product Details Navigation', 9, 'verify_element_visible', 'class', 'inventory_details_name', '', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 10, 'verify_element_visible', 'class', 'inventory_details_price', '', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 11, 'verify_element_visible', 'class', 'inventory_details_desc', '', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 12, 'take_screenshot', '', '', 'TC011_product_details', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 13, 'click', 'css', '.inventory_details_back_button', '', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 14, 'wait', '', '', '2', '', 'Yes'],
        ['TC011', 'Product Details Navigation', 15, 'verify_url', '', '', '', 'inventory.html', 'Yes'],

        # ──── TC012: Checkout Without Info (Negative) ────
        ['TC012', 'Checkout Without Info', 1, 'navigate', '', '', BASE, '', 'Yes'],
        ['TC012', 'Checkout Without Info', 2, 'type', 'id', 'user-name', USER, '', 'Yes'],
        ['TC012', 'Checkout Without Info', 3, 'type', 'id', 'password', PASS, '', 'Yes'],
        ['TC012', 'Checkout Without Info', 4, 'click', 'id', 'login-button', '', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 5, 'wait', '', '', '2', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 6, 'click', 'css', '.inventory_item:nth-child(1) .btn_inventory', '', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 7, 'wait', '', '', '2', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 8, 'click', 'css', '.shopping_cart_link', '', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 9, 'wait', '', '', '2', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 10, 'click', 'css', '.checkout_button', '', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 11, 'wait', '', '', '2', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 12, 'click', 'css', '.cart_button', '', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 13, 'wait', '', '', '2', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 14, 'verify_element_visible', 'css', 'h3[data-test="error"]', '', '', 'Yes'],
        ['TC012', 'Checkout Without Info', 15, 'verify_text', 'css', 'h3[data-test="error"]', '', 'First Name is required', 'Yes'],
        ['TC012', 'Checkout Without Info', 16, 'take_screenshot', '', '', 'TC012_checkout_error', '', 'Yes'],
    ]

    # Color-code rows by test case for readability
    tc_colors = [
        'FFFFFF', 'F2F7FB', 'FFFFFF', 'F2F7FB',
        'FFFFFF', 'F2F7FB', 'FFFFFF', 'F2F7FB',
        'FFFFFF', 'F2F7FB', 'FFFFFF', 'F2F7FB',
    ]
    tc_ids = list(dict.fromkeys(row[0] for row in sample_data))
    tc_color_map = {tc_id: tc_colors[i % len(tc_colors)] for i, tc_id in enumerate(tc_ids)}

    for row_idx, row_data in enumerate(sample_data, 2):
        fill_color = tc_color_map.get(row_data[0], 'FFFFFF')
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            if fill_color != 'FFFFFF':
                cell.fill = row_fill

    col_widths = [15, 30, 10, 25, 15, 45, 35, 25, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    # ═══════════════════════════════════════════
    # Sheet 3: Action Keywords Reference
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet('Action Keywords Reference')

    ref_headers = ['Action Keyword', 'Description', 'Locator_Type Required', 'Test_Data Usage', 'Example']
    for col, header in enumerate(ref_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    keyword_ref = [
        ['navigate / go_to_url', 'Open a URL in browser', 'No', 'URL or ${base_url}', '${base_url}'],
        ['click', 'Click on an element', 'Yes', 'Not used', ''],
        ['type / send_keys', 'Type text into an input field', 'Yes', 'Text to type or ${variable}', '${username}'],
        ['clear', 'Clear an input field', 'Yes', 'Not used', ''],
        ['select_dropdown', 'Select option from dropdown', 'Yes', 'Visible text of option', 'Price (low to high)'],
        ['verify_text', 'Verify element text contains value', 'Yes', 'Not used (use Expected_Result)', 'Products'],
        ['verify_title', 'Verify page title contains text', 'No', 'Expected title text', 'Swag Labs'],
        ['verify_url', 'Verify current URL contains text', 'No', 'Not used (use Expected_Result)', 'inventory.html'],
        ['verify_element_present', 'Check if element exists in DOM', 'Yes', 'Not used', ''],
        ['verify_element_visible', 'Check if element is visible on screen', 'Yes', 'Not used', ''],
        ['hover / mouse_over', 'Hover mouse over an element', 'Yes', 'Not used', ''],
        ['double_click', 'Double-click an element', 'Yes', 'Not used', ''],
        ['right_click', 'Right-click (context menu) an element', 'Yes', 'Not used', ''],
        ['wait / sleep', 'Pause execution for N seconds', 'No', 'Seconds to wait', '3'],
        ['press_key', 'Press a keyboard key', 'No', 'Key name (ENTER, TAB, etc.)', 'ENTER'],
        ['scroll_down', 'Scroll page down by N pixels', 'No', 'Pixels (default 500)', '500'],
        ['scroll_up', 'Scroll page up by N pixels', 'No', 'Pixels (default 500)', '300'],
        ['scroll_to_element', 'Scroll until element is visible', 'Yes', 'Not used', ''],
        ['switch_frame', 'Switch to an iframe element', 'Yes', 'Not used', ''],
        ['switch_default', 'Switch back to main page content', 'No', 'Not used', ''],
        ['switch_window', 'Switch to another browser tab/window', 'No', 'Window index (0,1,2..)', '1'],
        ['accept_alert', 'Click OK on browser alert popup', 'No', 'Not used', ''],
        ['dismiss_alert', 'Click Cancel on browser alert popup', 'No', 'Not used', ''],
        ['take_screenshot', 'Capture and save a screenshot', 'No', 'Screenshot filename', 'login_page'],
        ['refresh', 'Refresh/reload the current page', 'No', 'Not used', ''],
        ['go_back', 'Navigate to previous page (browser back)', 'No', 'Not used', ''],
        ['upload_file', 'Upload a file via file input', 'Yes', 'Absolute file path', 'C:/files/test.pdf'],
        ['store_text', 'Save element text into a variable', 'Yes', 'Variable name', 'user_name'],
        ['execute_js', 'Execute custom JavaScript code', 'No', 'JS code to execute', 'return document.title'],
        ['api_get', 'Send HTTP GET request to an API', 'No', 'API URL', 'http://localhost:8000/api/users'],
        ['api_post', 'Send HTTP POST request (URL in Locator_Value)', 'No', 'JSON request body', '{"key": "value"}'],
        ['verify_api_status', 'Verify last API response status code', 'No', 'Not used (use Expected_Result)', '200'],
        ['close_browser', 'Close the browser window', 'No', 'Not used', ''],
    ]

    for row_idx, row_data in enumerate(keyword_ref, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    ref_widths = [28, 40, 22, 38, 32]
    for i, width in enumerate(ref_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = width
    ws2.freeze_panes = 'A2'

    # ═══════════════════════════════════════════
    # Sheet 4: Locator Types Reference
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet('Locator Types Reference')

    loc_headers = ['Locator_Type', 'Description', 'Example Locator_Value', 'SauceDemo Example']
    for col, header in enumerate(loc_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    locator_ref = [
        ['id', 'Find element by HTML id attribute', 'login-button', 'id="user-name" / id="password" / id="login-button"'],
        ['name', 'Find element by HTML name attribute', 'username', 'name="user-name"'],
        ['xpath', 'Find element using XPath expression', '//button[@type="submit"]', '//div[@class="inventory_item"]'],
        ['css', 'Find element using CSS selector', 'div.container > button', '.inventory_item .btn_inventory'],
        ['class', 'Find element by CSS class name', 'btn-primary', 'inventory_list / product_sort_container'],
        ['tag', 'Find element by HTML tag name', 'input', 'img / button / select'],
        ['link_text', 'Find hyperlink by exact visible text', 'Click Here', 'Sauce Labs Backpack'],
        ['partial_link_text', 'Find hyperlink by partial visible text', 'Click', 'Sauce Labs'],
    ]

    for row_idx, row_data in enumerate(locator_ref, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    for i, width in enumerate([18, 38, 35, 48], 1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = width
    ws3.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"[OK] Test case template created: {output_path}")
    print(f"     Sheets: Test Plan | Test Cases (12 scenarios) | Action Keywords Reference | Locator Types Reference")
    return output_path


if __name__ == '__main__':
    create_test_case_template()
