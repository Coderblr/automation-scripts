"""Generate OrangeHRM test cases Excel file with verified selectors."""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = os.path.join('test_data', 'test_cases.xlsx')
os.makedirs('test_data', exist_ok=True)

wb = Workbook()
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

# ── Sheet 1: Test Plan ──
ws = wb.active
ws.title = 'Test Plan'
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 35

title_font = Font(bold=True, size=16, color='1F4E79')
section_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
section_font = Font(bold=True, size=12, color='FFFFFF')
label_font = Font(bold=True, size=11)
value_font = Font(size=11)
light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

ws.merge_cells('A1:D1')
c = ws.cell(1, 1, 'AUTOMATION TEST PLAN')
c.font = title_font
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:D2')
c = ws.cell(2, 1, 'OrangeHRM Web Application - Regression Test Suite')
c.font = Font(size=12, italic=True, color='4472C4')
c.alignment = Alignment(horizontal='center')

row = 4
for col in range(1, 5):
    c = ws.cell(row, col, 'PROJECT INFORMATION' if col == 1 else '')
    c.fill = section_fill
    c.font = section_font
    c.border = thin
ws.merge_cells(f'A{row}:D{row}')

info = [
    ('Title', 'OrangeHRM Automation Test Suite', 'Project ID', 'OHR-ATS-001'),
    ('Prepared By', 'QA Team', 'Date', datetime.now().strftime('%Y-%m-%d')),
    ('Tools Used', 'Python Selenium / Java Selenium', 'Version', '1.0'),
    ('Environment', 'Demo (opensource-demo.orangehrmlive.com)', 'Browser', 'Chrome / Edge / Firefox'),
    ('Test Type', 'Regression, Smoke, Functional', 'Priority', 'High'),
    ('Credentials', 'Admin / admin123', 'Status', 'Active'),
]
for i, (l1, v1, l2, v2) in enumerate(info):
    r = row + 1 + i
    ws.cell(r, 1, l1).font = label_font
    ws.cell(r, 1).fill = light_fill
    ws.cell(r, 1).border = thin
    ws.cell(r, 2, v1).font = value_font
    ws.cell(r, 2).border = thin
    ws.cell(r, 3, l2).font = label_font
    ws.cell(r, 3).fill = light_fill
    ws.cell(r, 3).border = thin
    ws.cell(r, 4, v2).font = value_font
    ws.cell(r, 4).border = thin

# ── Sheet 2: Test Cases ──
ws2 = wb.create_sheet('Test Cases')
headers = ['Test_Case_ID', 'Test_Case_Name', 'Step_No', 'Action',
           'Locator_Type', 'Locator_Value', 'Test_Data', 'Expected_Result', 'Run']

hdr_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True, size=11)
for col, h in enumerate(headers, 1):
    c = ws2.cell(1, col, h)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin

BASE = '${base_url}'
USER = '${username}'
PASS = '${password}'

data = [
    # TC001: Valid Login
    ['TC001', 'Valid Login', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC001', 'Valid Login', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC001', 'Valid Login', 3, 'verify_title', '', '', 'OrangeHRM', '', 'Yes'],
    ['TC001', 'Valid Login', 4, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC001', 'Valid Login', 5, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC001', 'Valid Login', 6, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC001', 'Valid Login', 7, 'wait', '', '', '4', '', 'Yes'],
    ['TC001', 'Valid Login', 8, 'verify_url', '', '', '', 'dashboard', 'Yes'],
    ['TC001', 'Valid Login', 9, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'Dashboard', 'Yes'],
    ['TC001', 'Valid Login', 10, 'take_screenshot', '', '', 'TC001_login_success', '', 'Yes'],

    # TC002: Invalid Login
    ['TC002', 'Invalid Login', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC002', 'Invalid Login', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC002', 'Invalid Login', 3, 'type', 'name', 'username', 'InvalidUser', '', 'Yes'],
    ['TC002', 'Invalid Login', 4, 'type', 'name', 'password', 'wrongpass', '', 'Yes'],
    ['TC002', 'Invalid Login', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC002', 'Invalid Login', 6, 'wait', '', '', '3', '', 'Yes'],
    ['TC002', 'Invalid Login', 7, 'verify_element_visible', 'css', '.oxd-alert-content-text', '', '', 'Yes'],
    ['TC002', 'Invalid Login', 8, 'verify_text', 'css', '.oxd-alert-content-text', '', 'Invalid credentials', 'Yes'],
    ['TC002', 'Invalid Login', 9, 'take_screenshot', '', '', 'TC002_invalid_login', '', 'Yes'],

    # TC003: Dashboard Verification
    ['TC003', 'Dashboard Verification', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC003', 'Dashboard Verification', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC003', 'Dashboard Verification', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC003', 'Dashboard Verification', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 7, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'Dashboard', 'Yes'],
    ['TC003', 'Dashboard Verification', 8, 'verify_element_visible', 'css', '.oxd-userdropdown-tab', '', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 9, 'verify_element_visible', 'xpath', "//span[text()='Admin']", '', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 10, 'verify_element_visible', 'xpath', "//span[text()='PIM']", '', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 11, 'verify_element_present', 'xpath', "//span[text()='Dashboard']", '', '', 'Yes'],
    ['TC003', 'Dashboard Verification', 12, 'take_screenshot', '', '', 'TC003_dashboard', '', 'Yes'],

    # TC004: Navigate to PIM Module
    ['TC004', 'Navigate to PIM Module', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 7, 'click', 'xpath', "//span[text()='PIM']", '', '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 8, 'wait', '', '', '3', '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 9, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'PIM', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 10, 'verify_url', '', '', '', 'pim', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 11, 'verify_element_visible', 'css', '.oxd-table', '', '', 'Yes'],
    ['TC004', 'Navigate to PIM Module', 12, 'take_screenshot', '', '', 'TC004_pim_module', '', 'Yes'],

    # TC005: Navigate to Admin Module
    ['TC005', 'Navigate to Admin Module', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 7, 'click', 'xpath', "//span[text()='Admin']", '', '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 8, 'wait', '', '', '3', '', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 9, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'Admin', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 10, 'verify_url', '', '', '', 'admin', 'Yes'],
    ['TC005', 'Navigate to Admin Module', 11, 'take_screenshot', '', '', 'TC005_admin_module', '', 'Yes'],

    # TC006: Navigate to My Info
    ['TC006', 'Navigate to My Info', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC006', 'Navigate to My Info', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC006', 'Navigate to My Info', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC006', 'Navigate to My Info', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC006', 'Navigate to My Info', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC006', 'Navigate to My Info', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC006', 'Navigate to My Info', 7, 'click', 'xpath', "//span[text()='My Info']", '', '', 'Yes'],
    ['TC006', 'Navigate to My Info', 8, 'wait', '', '', '3', '', 'Yes'],
    ['TC006', 'Navigate to My Info', 9, 'verify_url', '', '', '', 'pim', 'Yes'],
    ['TC006', 'Navigate to My Info', 10, 'take_screenshot', '', '', 'TC006_my_info', '', 'Yes'],

    # TC007: Navigate to Leave Module
    ['TC007', 'Navigate to Leave Module', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 7, 'scroll_to_element', 'xpath', "//span[text()='Leave']", '', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 8, 'wait', '', '', '1', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 9, 'click', 'xpath', "//span[text()='Leave']", '', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 10, 'wait', '', '', '3', '', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 11, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'Leave', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 12, 'verify_url', '', '', '', 'leave', 'Yes'],
    ['TC007', 'Navigate to Leave Module', 13, 'take_screenshot', '', '', 'TC007_leave_module', '', 'Yes'],

    # TC008: Navigate to Recruitment
    ['TC008', 'Navigate to Recruitment', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 7, 'click', 'xpath', "//span[text()='Recruitment']", '', '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 8, 'wait', '', '', '3', '', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 9, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'Recruitment', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 10, 'verify_url', '', '', '', 'recruitment', 'Yes'],
    ['TC008', 'Navigate to Recruitment', 11, 'take_screenshot', '', '', 'TC008_recruitment', '', 'Yes'],

    # TC009: Navigate to Directory
    ['TC009', 'Navigate to Directory', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC009', 'Navigate to Directory', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC009', 'Navigate to Directory', 3, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC009', 'Navigate to Directory', 4, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC009', 'Navigate to Directory', 5, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC009', 'Navigate to Directory', 6, 'wait', '', '', '4', '', 'Yes'],
    ['TC009', 'Navigate to Directory', 7, 'click', 'xpath', "//span[text()='Directory']", '', '', 'Yes'],
    ['TC009', 'Navigate to Directory', 8, 'wait', '', '', '3', '', 'Yes'],
    ['TC009', 'Navigate to Directory', 9, 'verify_text', 'css', '.oxd-topbar-header-breadcrumb h6', '', 'Directory', 'Yes'],
    ['TC009', 'Navigate to Directory', 10, 'verify_url', '', '', '', 'directory', 'Yes'],
    ['TC009', 'Navigate to Directory', 11, 'take_screenshot', '', '', 'TC009_directory', '', 'Yes'],

    # TC010: Logout
    ['TC010', 'Logout', 1, 'navigate', '', '', BASE, '', 'Yes'],
    ['TC010', 'Logout', 2, 'wait', '', '', '3', '', 'Yes'],
    ['TC010', 'Logout', 3, 'refresh', '', '', '', '', 'Yes'],
    ['TC010', 'Logout', 4, 'wait', '', '', '3', '', 'Yes'],
    ['TC010', 'Logout', 5, 'type', 'name', 'username', USER, '', 'Yes'],
    ['TC010', 'Logout', 6, 'type', 'name', 'password', PASS, '', 'Yes'],
    ['TC010', 'Logout', 7, 'click', 'css', 'button[type=submit]', '', '', 'Yes'],
    ['TC010', 'Logout', 8, 'wait', '', '', '5', '', 'Yes'],
    ['TC010', 'Logout', 9, 'verify_url', '', '', '', 'dashboard', 'Yes'],
    ['TC010', 'Logout', 10, 'click', 'css', '.oxd-userdropdown-tab', '', '', 'Yes'],
    ['TC010', 'Logout', 11, 'wait', '', '', '2', '', 'Yes'],
    ['TC010', 'Logout', 12, 'click', 'xpath', "//a[text()='Logout']", '', '', 'Yes'],
    ['TC010', 'Logout', 13, 'wait', '', '', '3', '', 'Yes'],
    ['TC010', 'Logout', 14, 'verify_url', '', '', '', 'login', 'Yes'],
    ['TC010', 'Logout', 15, 'verify_element_visible', 'name', 'username', '', '', 'Yes'],
    ['TC010', 'Logout', 16, 'take_screenshot', '', '', 'TC010_logout_success', '', 'Yes'],
]

tc_colors = ['FFFFFF', 'F2F7FB']
tc_ids = list(dict.fromkeys(r[0] for r in data))
tc_map = {tc: tc_colors[i % 2] for i, tc in enumerate(tc_ids)}

for ri, rd in enumerate(data, 2):
    color = tc_map.get(rd[0], 'FFFFFF')
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for ci, val in enumerate(rd, 1):
        c = ws2.cell(ri, ci, val)
        c.border = thin
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        if color != 'FFFFFF':
            c.fill = fill

widths = [15, 30, 10, 25, 15, 45, 35, 25, 8]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w
ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = 'A2'

wb.save(OUTPUT)
print(f'[OK] OrangeHRM test cases created: {OUTPUT}')
print(f'     {len(tc_ids)} test cases, {len(data)} total steps')
